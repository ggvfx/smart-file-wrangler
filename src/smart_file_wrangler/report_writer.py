"""
report_writer.py

Report output generation for Smart File Wrangler.

Purpose:
- Consumes metadata rows prepared upstream
- Writes CSV, JSON, Excel, and folder tree text outputs
- Converts absolute paths to relative paths for reports
- Provides sorting of report rows for stable output order

Terminology:
- metadata rows = dictionaries describing discovered files or sequences
- folder tree = a printed view of the organised folder structure on disk
- report items = the list of logical units passed from pipeline orchestration
"""

import csv
import json
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .media_item import MediaItem

# ----------------------------------------------------------------------
# Sorting helpers
# ----------------------------------------------------------------------
# --- Sorting is for stable report output order, not media filtering

media_type_order = {
    "video": 0,
    "image": 1,
    "audio": 2,
    "other": 3,
}

# ----------------------------------------------------------------------
# Path and filename helpers
# ----------------------------------------------------------------------
# --- Relative paths are used for reports to match organised folder tree

def _make_relative_path(path_string: str, root_folder: Path | str) -> str:
    """
    Convert an absolute path string to a path relative to the input folder.

    Args:
        path_string (str): Original absolute file path string
        root_folder (Path | str): Folder to make the path relative to

    Returns:
        str: A relative path string if possible, otherwise the original path

    Notes:
        - This is a pure string transformation, no filesystem scanning happens here.
        - Errors fall back safely to original input to preserve behavior.
    """
    if not path_string:
        return ""

    try:
        return str(
            Path(path_string)
            .resolve()
            .relative_to(Path(root_folder).resolve())
        )
    except Exception:
        return path_string


def _make_sequence_filename(metadata):
    """
    Get the filename for a frame sequence from its metadata row.

    Args:
        metadata (dict): A metadata dictionary produced upstream

    Returns:
        str: The pre-constructed sequence filename if present, otherwise empty string

    Notes:
        - Frame sequence filenames are created upstream and reused here.
        - This function does not infer or scan for files.
    """
    return metadata.get("filename", "")

# ----------------------------------------------------------------------
# Shared row preparation
# ----------------------------------------------------------------------

def _prepare_report_row(row, root_folder):
    """
    Prepare a metadata row for report output.

    This function:
    - Computes relative file_path
    - Computes filename (sequence vs normal file)
    - Adds human-readable file_size
    - Leaves all other fields unchanged

    Returns:
        dict: Prepared row (copy)
    """
    output_row = dict(row)

    original_path = row.get("file_path", "")
    relative_path = _make_relative_path(original_path, root_folder)

    # filename logic
    if row.get("frame_count"):
        filename = _make_sequence_filename(row)
    else:
        filename = Path(original_path).name if original_path else ""

    output_row["filename"] = filename
    output_row["file_path"] = relative_path

    # human-readable file size
    size_bytes = row.get("file_size_bytes")
    output_row["file_size"] = _format_file_size(size_bytes)

    return output_row

# ----------------------------------------------------------------------
# CSV report
# ----------------------------------------------------------------------

def write_csv_report(data, output_path, root_folder, config):
    # unwrap MediaItem internally, no behavior change to legacy callers
    data = [i.sequence_info if isinstance(i, MediaItem) else i for i in data]
    """
    Write metadata rows to CSV.

    Args:
        data (list[Path | dict]): Legacy report items from pipeline
        output_path (Path | str): Where to write the CSV
        root_folder (Path | str): Root input folder for relative paths
        config (Config): Config object providing metadata field list

    Behavior:
        - Header row is always written first
        - filename column is first for readability
        - file paths are made relative internally
        - file sizes are human-readable strings, not raw bytes

    Notes:
        - This function does not scan the filesystem
        - It preserves legacy behavior and data shape
    """

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    selected_fields = list(config.metadata_fields)

    if "file_path" not in selected_fields:
        selected_fields.insert(0, "file_path")

    header = ["filename"] + [field_name for field_name in selected_fields if field_name != "filename"]

    with output_path.open("w", newline="", encoding="utf-8") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=header)
        writer.writeheader()

        for row in data:
            prepared = _prepare_report_row(row, root_folder)

            # CSV expects file_size, not raw bytes
            prepared = {
                field_name: prepared.get(field_name, "")
                for field_name in header
            }

            writer.writerow(prepared)

    print(f'CSV report written: "{output_path}" ({len(data)} rows)')

# ----------------------------------------------------------------------
# JSON report
# ----------------------------------------------------------------------

def write_json_report(data, output_path, root_folder, config):
    # unwrap MediaItem internally
    data = [i.sequence_info if isinstance(i, MediaItem) else i for i in data]
    """
    Write report data to a JSON file.

    Matches CSV output:
    - filename field
    - relative file_path
    - human-readable file_size
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    output_rows = []

    for row in data:
        prepared = _prepare_report_row(row, root_folder)

        # Remove raw byte field for JSON output
        prepared.pop("file_size_bytes", None)

        filtered = {
            field_name: prepared.get(field_name)
            for field_name in ["filename"] + config.metadata_fields
            if field_name in prepared
        }

        output_rows.append(filtered)

    with output_path.open("w", encoding="utf-8") as output_file:
        json.dump(output_rows, output_file, indent=2)

    print(f'JSON report written: "{output_path}" ({len(output_rows)} items)')

# ----------------------------------------------------------------------
# Folder tree output
# ----------------------------------------------------------------------
# --- Tree building uses metadata rows, not raw filesystem walks

def _build_tree(items):
    """
    Build a nested dictionary representing folder structure from relative metadata rows.

    Args:
        items (list[dict]): Metadata rows containing relative file_path strings

    Returns:
        dict: Nested mapping of folder → file → None

    Notes:
        - This builds a logical tree from metadata, not a disk walk
        - The disk folder tree is generated separately upstream
    """
    tree = defaultdict(dict)

    for item in items:
        path = Path(item["file_path"])
        parts = path.parts

        current = tree
        for part in parts[:-1]:
            current = current.setdefault(part, {})

        current[parts[-1]] = None

    return tree


def _write_tree_lines(tree, indent=""):
    lines = []

    entries = sorted(tree.items())

    for index, (name, subtree) in enumerate(entries):
        is_last = index == len(entries) - 1
        branch = "└─ " if is_last else "├─ "
        lines.append(f"{indent}{branch}{name}")

        if isinstance(subtree, dict):
            extension = "   " if is_last else "│  "
            lines.extend(
                _write_tree_lines(subtree, indent + extension)
            )

    return lines


def write_folder_tree(items, output_path, root_folder, config):
    """
    Write a text-based folder tree.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    root_folder = Path(root_folder)
    root_name = root_folder.name
    thumb_folder = config.thumb_folder_name

    relative_items = []

    for item in items:
        relative_path = Path(item["file_path"]).relative_to(root_folder)

        # Skip thumbnail folders entirely
        if thumb_folder in relative_path.parts:
            continue

        new_item = dict(item)
        new_item["file_path"] = str(relative_path)
        relative_items.append(new_item)

    tree = _build_tree(relative_items)
    lines = _write_tree_lines(tree)
    lines = [root_name] + lines

    with output_path.open("w", encoding="utf-8") as tree_file:
        for line in lines:
            tree_file.write(line + "\n")

    print(f'Folder tree written: "{output_path}"')

# ----------------------------------------------------------------------
# Excel report
# ----------------------------------------------------------------------

def write_excel_report(data, output_path, root_folder, config):
    # unwrap MediaItem internally
    data = [i.sequence_info if isinstance(i, MediaItem) else i for i in data]
    """
    Write report data to an Excel (.xlsx) file.

    Mirrors CSV/JSON output:
    - filename first
    - relative file_path
    - human-readable file_size
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    selected_fields = list(config.metadata_fields)

    if "file_path" not in selected_fields:
        selected_fields.insert(0, "file_path")

    selected_fields = [
        "file_size" if field_name == "file_size_bytes" else field_name
        for field_name in selected_fields
    ]

    header = [
        "filename"
    ] + [
        field_name
        for field_name in selected_fields
        if field_name != "filename"
    ]
    sheet.append(header)

    for row in data:
        prepared = _prepare_report_row(row, root_folder)
        sheet.append([prepared.get(col, "") for col in header])

    # Auto-size columns
    for col_index, column in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[
            get_column_letter(col_index)
        ].width = max_length + 2

    workbook.save(output_path)
    print(f'Excel report written: "{output_path}" ({len(data)} rows)')

# ----------------------------------------------------------------------
# Sorting
# ----------------------------------------------------------------------

def sort_report_items(data, root_folder):
    """
    Sort report items by:
    - folder depth
    - parent folder
    - filename
    - media type
    - extension
    """
    def sort_key(item):
        rel_path = Path(
            _make_relative_path(item.get("file_path", ""), root_folder)
        )

        depth = len(rel_path.parts) - 1
        parent = rel_path.parent.as_posix()

        if isinstance(item, MediaItem) and item.kind == "sequence":
            filename = _make_sequence_filename(item)
        else:
            filename = Path(item.sequence_info.get("file_path", "")).name if isinstance(item, MediaItem) else Path(item.get("file_path", "")).name

        media_type = item.get("media_type", "other")
        media_order = media_type_order.get(media_type, 99)
        extension = item.get("extension", "")

        return (depth, parent, filename.lower(), media_order, extension.lower())

    return sorted(data, key=sort_key)

# ----------------------------------------------------------------------
# Public entry point
# ----------------------------------------------------------------------

def generate_reports(
    metadata,
    input_folder,
    output_dir,
    fields,
    sort_by=None,
    reverse=False,
    csv_enabled=False,
    json_enabled=False,
    tree_enabled=False,
    excel_enabled=False,
    config=None,
):
    if config is None:
        raise ValueError(
            "generate_reports() now requires a Config object. "
            "Pipeline must pass config explicitly."
        )

    # NOTE: sort_by/reverse preserved for API compatibility
    sorted_data = sort_report_items(metadata, input_folder)

    if csv_enabled:
        write_csv_report(sorted_data, Path(output_dir) / "report.csv", root_folder=input_folder, config=config)

    if json_enabled:
        write_json_report(sorted_data, Path(output_dir) / "report.json", root_folder=input_folder, config=config)

    if tree_enabled:
        write_folder_tree(sorted_data, Path(output_dir) / "folder_tree.txt", root_folder=input_folder, config=config)

    if excel_enabled:
        write_excel_report(sorted_data, Path(output_dir) / "report.xlsx", root_folder=input_folder, config=config)

# ----------------------------------------------------------------------
# Formatting helpers
# ----------------------------------------------------------------------

def _format_file_size(bytes_value):
    """
    Convert a byte count to a human-readable string.
    """
    if bytes_value is None:
        return ""

    bytes_value = int(bytes_value)

    if bytes_value < 1024:
        return f"{bytes_value} B"

    kb = bytes_value / 1024
    if kb < 1024:
        return f"{kb:.1f} KB"

    mb = kb / 1024
    if mb < 1024:
        return f"{mb:.2f} MB"

    gb = mb / 1024
    return f"{gb:.2f} GB"
